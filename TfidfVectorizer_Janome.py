import numpy as np
import openpyxl
import pandas as pd
from janome.tokenizer import Tokenizer
from sklearn.feature_extraction.text import TfidfVectorizer

# 設定
excel_file = "question_list.xlsx"

# 1. Excelデータの読み込み
# 1行目からデータとして読み込むため、header=Noneを指定
df_excel = pd.read_excel(excel_file, sheet_name="Sheet1", header=None)

# 列を分解 (1列目: インデックス, 2列目: 質問文)
indices = df_excel[0].tolist()
questions = df_excel[1].fillna("").astype(str).tolist()  # 空白（NaN）対策

# 2. わかち書き処理（一括で1回だけ実行）
t = Tokenizer()
wakati_lists = []
wakati_sentences = []

for q in questions:
    # 形態素解析して表面形（単語）のリストを作成
    tokens = [token.surface for token in t.tokenize(q)]
    wakati_lists.append(tokens)
    # TfidfVectorizer用にスペース区切りの文字列も用意
    wakati_sentences.append(" ".join(tokens))

# 3. ベクトル化 (TF-IDF)
# token_pattern=r"(?u)\b\w+\b" は1文字の単語も含めるための設定
vectorizer = TfidfVectorizer(
    analyzer="word", token_pattern=r"(?u)\b\w+\b", binary=True, use_idf=True
)
vecs = vectorizer.fit_transform(wakati_sentences).toarray()

# 全単語のリスト
headers = list(vectorizer.get_feature_names_out())

# 4. 元のExcelファイルへの書き込み
wb = openpyxl.load_workbook(excel_file)
ws = wb["Sheet1"]

for i, (w_list, vec) in enumerate(zip(wakati_lists, vecs), start=1):
    ws.cell(row=i, column=3).value = str(w_list)  # 3列目: わかち書き
    ws.cell(row=i, column=4).value = str(vec.tolist())  # 4列目: ベクトル
    print(f"Excel書き込み進捗: {i} / {len(questions)}")

# 5列目の1行目に全単語リストを記録
ws.cell(row=1, column=5).value = str(headers)
wb.save(excel_file)

# 5. CSVへの出力 (Pandas DataFrame)
df_csv = pd.DataFrame(data=vecs, index=indices, columns=headers)
df_csv.to_csv("detaset.csv", encoding="utf-8-sig")

# 出力確認
print("\n--- 処理が正常に完了しました ---")
np.set_printoptions(threshold=np.inf)
print(f"全単語数: {len(headers)}")
print(df_csv)